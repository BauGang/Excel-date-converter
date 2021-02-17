import openpyxl
from pathlib import Path
import datetime
from datetime import datetime, timedelta
data = datetime.now()
giorno = data.strftime("%Y-%m-%d 00:00:00")
xlsx_file = Path('add_folder_here', 'add_file_here')
wb_obj = openpyxl.load_workbook(xlsx_file)
ws = wb_obj.active

HoursOfWork = 7

def convertDate(ordinal):
    epochStart = datetime(1899, 12, 31)
    if ordinal is not None:
        if ordinal >= 60:
            ordinal -= 1
        return epochStart + timedelta(days=ordinal)

row_count = ws.max_row
count = 0

for j in range(1, row_count, HoursOfWork):  # skip the hours, so he can take the day 
    a = ws.cell(row=j, column=1).value
    b = convertDate(a)
    if giorno == str(b):
        for x in range(j+1, row_count):     # you can put ws.max_row but remember, the library take the last row changed,  
            if count == HoursOfWork:        # so if you had 300 rows but you edited the row 1000 too the librarie take the row 1000.
                break
            else:
                print(ws.cell(row=x,column=2).value)
                count += 1
